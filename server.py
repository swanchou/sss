import re
import io
import json
import os
import sys
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook


def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_dist_path():
    base = get_base_path()
    dist_path = os.path.join(base, 'dist')
    if os.path.exists(dist_path):
        return dist_path
    return os.path.join(base, 'dist')


app = Flask(__name__, static_folder=get_dist_path(), static_url_path='')
CORS(app)

MOBILE_PATTERNS = [
    re.compile(r'(?<![a-zA-Z0-9])(?:\+?86[-\s]?)?1[3-9]\d{9}(?!\d)'),
    re.compile(r'(?<![a-zA-Z0-9])1[3-9]\d[-\s]?\d{4}[-\s]?\d{4}(?!\d)'),
    re.compile(r'(?<![a-zA-Z0-9])(?:\+?86[-\s]?)?1[3-9]\d\s\d{4}\s\d{4}(?!\d)'),
]

SN_PATTERN = re.compile(r'(?<![a-zA-Z0-9])[A-Za-z0-9]{16}(?![a-zA-Z0-9])')

RISK_PATTERN = re.compile(r'(华为账号|来电号码|备用号码)\s*[:：]?\s*1[3-9]\d{9}')
SN_RISK_PATTERN = re.compile(r'(sn|SN|序列号)\s*[:：]\s*[A-Za-z0-9]{16}')


def load_trusted_phones():
    base = get_base_path()
    json_path = os.path.join(base, 'src', '华为授权服务中心_整理.json')
    if not os.path.exists(json_path):
        json_path = os.path.join(base, '华为授权服务中心_整理.json')
    phones = set()
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for item in data:
            raw = item.get('phone', '')
            normalized = re.sub(r'[\s\-\(\)\+\.]', '', raw).lstrip('86')
            if len(normalized) == 11 and re.match(r'^1[3-9]\d{9}$', normalized):
                phones.add(normalized)
            elif normalized:
                phones.add(normalized)
    except Exception as e:
        print(f'加载号码库失败: {e}')
    return phones

TRUSTED_PHONES = load_trusted_phones()


def normalize_phone(raw):
    return re.sub(r'[\s\-\(\)\+\.]', '', raw).lstrip('86')


def is_valid_mobile(digits):
    return len(digits) == 11 and bool(re.match(r'^1[3-9]\d{9}$', digits))


def extract_phones(text):
    if text is None:
        return []
    s = str(text).strip()
    if not s:
        return []
    s = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', s)
    phones = set()
    for pattern in MOBILE_PATTERNS:
        for match in pattern.finditer(s):
            normalized = normalize_phone(match.group())
            if is_valid_mobile(normalized):
                phones.add(normalized)
    return list(phones)


def extract_sns(text):
    if text is None:
        return []
    s = str(text).strip()
    if not s:
        return []
    s = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', s)
    sns = set()
    for match in SN_PATTERN.finditer(s):
        matched = match.group()
        has_letter = bool(re.search(r'[a-zA-Z]', matched))
        has_digit = bool(re.search(r'[0-9]', matched))
        print(f'[SN匹配] 原文: "{s[:80]}" | 候选: "{matched}" | 含字母: {has_letter} | 含数字: {has_digit}')
        if has_letter and has_digit:
            print(f'  -> 确认为SN (字母+数字)')
            sns.add(matched)
        else:
            print(f'  -> 排除 (缺少{"数字" if not has_digit else "字母"})')
    return list(sns)


def classify_risk(text, phones, sns):
    if RISK_PATTERN.search(text):
        print(f'[风险判定] 重危(手机号关键词): "{text[:50]}..."')
        return '重危'
    if SN_RISK_PATTERN.search(text):
        print(f'[风险判定] 重危(SN关键词): "{text[:50]}..."')
        return '重危'
    for p in phones:
        if p in TRUSTED_PHONES:
            print(f'[风险判定] 低风险(号码库匹配): "{text[:50]}..."')
            return '低风险'
    print(f'[风险判定] 普通: "{text[:50]}..."')
    return '普通'


def parse_col_ref(ref):
    ref = ref.strip().upper()
    col_idx = 0
    for ch in ref:
        if 'A' <= ch <= 'Z':
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
        else:
            break
    return col_idx


def col_letter(col_idx):
    result = ''
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


@app.route('/')
def serve_index():
    dist_path = get_dist_path()
    return send_from_directory(dist_path, 'index.html')


@app.route('/<path:path>')
def serve_static(path):
    dist_path = get_dist_path()
    file_path = os.path.join(dist_path, path)
    if os.path.exists(file_path):
        return send_from_directory(dist_path, path)
    return send_from_directory(dist_path, 'index.html')


@app.route('/api/extract', methods=['POST'])
def extract():
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '文件名为空'}), 400
    valid_exts = ('.xlsx', '.xls', '.csv')
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if f'.{ext}' not in valid_exts:
        return jsonify({'error': '不支持的文件格式，请上传 .xlsx / .xls / .csv'}), 400

    target_col = request.form.get('target_col', '').strip()
    display_col = request.form.get('display_col', '').strip()
    if not target_col:
        return jsonify({'error': '请填写识别目标列号'}), 400
    if not display_col:
        return jsonify({'error': '请填写返回数据列号'}), 400

    target_idx = parse_col_ref(target_col)
    display_idx = parse_col_ref(display_col)
    if target_idx == 0 or display_idx == 0:
        return jsonify({'error': '列号格式不正确，请输入如 A、B、C 等'}), 400

    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400

    results = []
    riskCounts = {'普通': 0, '低风险': 0, '重危': 0}
    typeCounts = {'电话': 0, 'SN': 0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = max(target_idx, display_idx)
        for row in ws.iter_rows(min_row=1, max_col=max_col):
            row_num = row[0].row
            target_cell = row[target_idx - 1] if target_idx <= len(row) else None
            display_cell = row[display_idx - 1] if display_idx <= len(row) else None

            target_text = str(target_cell.value).strip() if target_cell and target_cell.value is not None else ''
            if not target_text:
                continue

            phones = extract_phones(target_text)
            sns = extract_sns(target_text)

            if not phones and not sns:
                continue

            display_text = str(display_cell.value).strip() if display_cell and display_cell.value is not None else ''
            risk_level = classify_risk(target_text, phones, sns)
            riskCounts[risk_level] += 1

            if phones:
                typeCounts['电话'] += 1
                results.append({
                    'displayValue': display_text,
                    'rowNum': row_num,
                    'phones': phones,
                    'sns': [],
                    'originalText': target_text,
                    'riskLevel': risk_level,
                    'dataType': '电话',
                })

            if sns:
                typeCounts['SN'] += 1
                results.append({
                    'displayValue': display_text,
                    'rowNum': row_num,
                    'phones': [],
                    'sns': sns,
                    'originalText': target_text,
                    'riskLevel': risk_level,
                    'dataType': 'SN',
                })

    wb.close()

    all_phones = set()
    all_sns = set()
    for r in results:
        for p in r['phones']:
            all_phones.add(p)
        for s in r['sns']:
            all_sns.add(s)

    return jsonify({
        'fileName': file.filename,
        'results': results,
        'totalPhones': len(all_phones),
        'totalSNs': len(all_sns),
        'targetCol': target_col.upper(),
        'displayCol': display_col.upper(),
        'riskCounts': riskCounts,
        'typeCounts': typeCounts,
    })


if __name__ == '__main__':
    is_frozen = getattr(sys, 'frozen', False)
    app.run(host='127.0.0.1', port=5001, debug=not is_frozen)
